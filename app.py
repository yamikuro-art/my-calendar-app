from flask import Flask, request, jsonify, render_template
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/save_to_excel', methods=['POST'])
def save_to_excel():
    try:
        data = request.get_json()
        print("Received data:", data)  # デバッグ用に受け取ったデータを出力

        if not data:
            raise ValueError("No data provided")

        df = pd.DataFrame(data)

        # 保存するディレクトリを確認または作成
        directory = 'output'
        if not os.path.exists(directory):
            os.makedirs(directory)

        # Excelファイルにデータを保存
        file_path = os.path.join(directory, 'calendar_data.xlsx')

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Calendar Data')
            workbook = writer.book
            worksheet = writer.sheets['Calendar Data']

            # フォントとアライメントのスタイルを設定
            header_font = Font(size=12, bold=True)
            cell_font = Font(size=10)
            alignment = Alignment(horizontal='center', vertical='center')
            fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 休日のセルに色をつける

            # ヘッダーのスタイルを適用
            for cell in worksheet[1]:
                cell.font = header_font
                cell.alignment = alignment

            # データセルのスタイルを適用
            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.font = cell_font
                    cell.alignment = alignment
                    if row[3].value == 'Yes':  # isHoliday列が'Yes'の場合
                        cell.fill = fill

        return jsonify({'success': True, 'message': 'Data successfully saved to Excel.'})
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)








